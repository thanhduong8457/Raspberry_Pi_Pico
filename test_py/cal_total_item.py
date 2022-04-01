#!/usr/bin/python
# caculate total number of item
# $Id: cal_total_item.py, v1.0 2022/02/17 $
# Creator: Tien Hoang
# Contributor: Duong Le
# ------------------------------------------------------------------
# Copyright(c) 2022 BanVien Co., Ltd.
# This program must be used solely for the purpose for which
# it was furnished by BanVien Corporation. No part of this
# program may be reproduced or disclosed to others, in any form,
# without the prior written permission of BanVien Corporation.
# ------------------------------------------------------------------

import sys
from xlrd import open_workbook
from openpyxl import load_workbook

# ******************************************
# Function     : DumpUsageMessage
# Description  : Dump usage messages
# Parameter    : None
# Return value : None
# ******************************************
def DumpUsageMessage():
    print("Wrong or missing argument(s)!                                                        ")
    print("=====================================================================================")
    print("= Usage: python cal_total_item.py <checklist_path>                                  =")
    print("=                <checklist_path>         : checklist                               =")
    print("=====================================================================================")

# ******************************************
# Function     : cal_total_item
# Description  : cal total item
# Parameter    : total rows and cols of a sheet
# Return value : first_Num_of_items, end_Num_of_items, first_Finished_items, end_Finished_items
# ******************************************
def cal_total_item(rows, cols):
    cols_Num_of_items   = get_num_of_items_column(rows,cols)
    cols_Finished_items = get_finished_items_column(rows,cols)
    rows_end_Total      = get_total_row(rows)

    first_Num_of_items = sheet_reg.cell_value(4, cols_Num_of_items)
    end_Num_of_items   = sheet_reg.cell_value(rows_end_Total  , cols_Num_of_items)

    first_Finished_items = sheet_reg.cell_value(4, cols_Finished_items)
    end_Finished_items   = sheet_reg.cell_value(rows_end_Total  , cols_Finished_items)
    
    print("[" + str(first_Num_of_items) + ", " + str(first_Finished_items) + "]")
    print("[" + str(end_Num_of_items) + ", " + str(end_Finished_items) + "]")
    
    return first_Num_of_items, end_Num_of_items, first_Finished_items, end_Finished_items

# ******************************************
# Function     : get_num_of_items_column
# Description  : Get position of Num. of items
# Parameter    : total rows and cols of a sheet
# Return value : item: position column of Num. of items
# ******************************************
def get_num_of_items_column(rows,cols):
    for column in range (0,cols):
        index = 0
        try:
            str_val = sheet_reg.cell_value(2,column)
            if "Num. of" in str(str_val):
                index = column
                break
        except:
            for row in range(0, rows):
                str_val = sheet_reg.cell_value(row,column)
                if "Num. of" in str(str_val):
                    index = column
                    break
    return index

# ******************************************
# Function     : get_finished_items_column
# Description  : Get position of Finished items
# Parameter    : total rows and cols of a sheet
# Return value : item: position column of Finished items
# ******************************************
def get_finished_items_column(rows,cols):
    for column in range (0,cols):
        index = 0
        try:
            str_val = sheet_reg.cell_value(2,column)
            if "Finished" in str(str_val):
                index = column
                break
        except:
            for row in range(0, rows):
                str_val = sheet_reg.cell_value(row,column)
                if "Finished" in str(str_val):
                    index = column
                    break
    return index

# ******************************************
# Function     : get_total_row
# Description  : Get position of Total
# Parameter    : total rows and cols of a sheet
# Return value : item: position row of Total
# ******************************************
def get_total_row(rows):
    rows_end_Total = False
    for row in range (0,rows):
        index_first = 0
        index_end = 0
        str_val = sheet_reg.cell_value(row,1)

        if ("Total" in str(str_val)):
            rows_end_Total = row
    return rows_end_Total

# ******************************************
# Function     : __main__
# Description  : main function
# ******************************************
if __name__ == '__main__':
    temp_total = 0
    total_first_Num_of_items = 0
    total_end_Num_of_items = 0
    total_first_Finished_items = 0
    total_end_Finished_items = 0
    if (len(sys.argv)>1):
        # Get path REQ file from argument 1
        path_file_REQ = sys.argv[1]
        # Get optional debug flag as argument 2
    else:
        # Dump usage message
        DumpUsageMessage()
        # Stop programing
        sys.exit(-1)
    try:
        # Define work book REQ
        wb_REQ = open_workbook(path_file_REQ)
        # print(path_file_REQ)
    except Exception:
        # Dump usage message
        print("Error: Cannot open " + path_file_REQ)
        # Stop programing
        sys.exit(-1)
    # Initialize Registers sheet in REQ

    workbook = load_workbook(filename=path_file_REQ)
    # print(f"Worksheet names: {workbook.sheetnames}")
    for sheets in workbook:
        print("")
        print(sheets)
        if sheets.sheet_state == "hidden": 
            print("=> this sheet is hidden")
            continue
        # Initialize Registers sheet in REQ
        sheet_reg = wb_REQ.sheet_by_name(sheets.title)
        first_Num_of_items, end_Num_of_items, first_Finished_items, end_Finished_items = cal_total_item(sheet_reg.nrows, sheet_reg.ncols)
        if first_Num_of_items != "":
            total_first_Num_of_items = total_first_Num_of_items + first_Num_of_items
        if end_Num_of_items != "":
            total_end_Num_of_items = total_end_Num_of_items + end_Num_of_items
        if first_Finished_items != "":
            total_first_Finished_items = total_first_Finished_items + first_Finished_items
        if end_Finished_items != "":
            total_end_Finished_items = total_end_Finished_items + end_Finished_items

    print("")
    print("============================== tienhoang: Hello team ========================================")
    print("[total_first_Num_of_items, total_first_Finished_items]    =   [" + str(int(total_first_Num_of_items)) + ", " + str(int(total_first_Finished_items)) + "]")
    print("[total_end_Num_of_items  , total_end_Finished_items  ]    =   [" + str(int(total_end_Num_of_items)) + ", " + str(int(total_end_Finished_items)) + "]")
    print("============================== tienhoang: Bye team ==========================================")

